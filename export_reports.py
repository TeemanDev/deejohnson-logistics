# export_reports.py - Excel export WITHOUT pandas
from flask import send_file, session, redirect, url_for, flash
from functools import wraps
from datetime import datetime
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

EXPORT_FOLDER = 'exports'
if not os.path.exists(EXPORT_FOLDER):
    os.makedirs(EXPORT_FOLDER)

def login_required_export(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            flash('Please login to access reports', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def create_excel_file(data, headers, sheet_name, filename):
    """Create Excel file using openpyxl directly"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3B5C", end_color="0F3B5C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_all_shipments(shipments):
    headers = [
        'Tracking Code', 'Customer Name', 'Customer Email', 'Customer Phone',
        'Origin', 'Destination', 'Package Type', 'Weight (kg)', 'Status',
        'Current Location', 'Estimated Delivery', 'Last Update', 'Direction'
    ]
    data = []
    for s in shipments:
        data.append([
            s.tracking_code,
            s.customer_name or '-',
            s.customer_email or '-',
            s.customer_phone or '-',
            s.origin or '-',
            s.destination or '-',
            s.package_type or '-',
            s.package_weight or 0,
            s.status[:50] if s.status else '-',
            s.current_location or '-',
            s.estimated_delivery or '-',
            s.last_update.strftime('%Y-%m-%d %H:%M') if s.last_update else '-',
            getattr(s, 'shipment_direction', 'outgoing')
        ])
    return create_excel_file(data, headers, 'All Shipments', 'all_shipments')

def export_outgoing_shipments(shipments):
    outgoing = [s for s in shipments if getattr(s, 'shipment_direction', 'outgoing') != 'incoming']
    headers = [
        'Tracking Code', 'Customer Name', 'Customer Email', 'Customer Phone',
        'Origin', 'Destination', 'Package Type', 'Weight (kg)', 'Status',
        'Current Location', 'Estimated Delivery', 'Partner Courier', 'Partner Tracking'
    ]
    data = []
    for s in outgoing:
        data.append([
            s.tracking_code,
            s.customer_name or '-',
            s.customer_email or '-',
            s.customer_phone or '-',
            s.origin or '-',
            s.destination or '-',
            s.package_type or '-',
            s.package_weight or 0,
            s.status[:50] if s.status else '-',
            s.current_location or '-',
            s.estimated_delivery or '-',
            s.partner_courier or '-',
            s.partner_tracking or '-'
        ])
    return create_excel_file(data, headers, 'Outgoing Shipments', 'outgoing_shipments')

def export_incoming_shipments(shipments):
    incoming = [s for s in shipments if getattr(s, 'shipment_direction', 'outgoing') == 'incoming']
    headers = [
        'Tracking Code', 'Customer Name', 'Customer Phone', 'Origin Country',
        'Origin City', 'Destination', 'Package Type', 'Weight (kg)', 'Status',
        'Original Courier', 'Original Tracking', 'Expected Arrival', 'Delivery Address'
    ]
    data = []
    for s in incoming:
        data.append([
            s.tracking_code,
            s.customer_name or '-',
            s.customer_phone or '-',
            getattr(s, 'origin_country', '-'),
            getattr(s, 'origin_city', '-'),
            s.destination or '-',
            s.package_type or '-',
            s.package_weight or 0,
            s.status[:50] if s.status else '-',
            getattr(s, 'partner_courier_original', '-'),
            getattr(s, 'partner_tracking_original', '-'),
            getattr(s, 'expected_arrival_date', '-'),
            getattr(s, 'delivery_address', '-')
        ])
    return create_excel_file(data, headers, 'Incoming Shipments', 'incoming_shipments')

def export_delivered_shipments(shipments):
    delivered = [s for s in shipments if 'Delivered' in s.status]
    headers = [
        'Tracking Code', 'Customer Name', 'Origin', 'Destination',
        'Package Type', 'Weight (kg)', 'Delivered Date', 'Direction'
    ]
    data = []
    for s in delivered:
        data.append([
            s.tracking_code,
            s.customer_name or '-',
            s.origin or '-',
            s.destination or '-',
            s.package_type or '-',
            s.package_weight or 0,
            s.last_update.strftime('%Y-%m-%d') if s.last_update else '-',
            getattr(s, 'shipment_direction', 'outgoing')
        ])
    return create_excel_file(data, headers, 'Delivered Shipments', 'delivered_shipments')

def export_pending_shipments(shipments):
    pending = [s for s in shipments if 'Delivered' not in s.status]
    headers = [
        'Tracking Code', 'Customer Name', 'Origin', 'Destination',
        'Package Type', 'Weight (kg)', 'Current Status', 'Direction'
    ]
    data = []
    for s in pending:
        data.append([
            s.tracking_code,
            s.customer_name or '-',
            s.origin or '-',
            s.destination or '-',
            s.package_type or '-',
            s.package_weight or 0,
            s.status[:50] if s.status else '-',
            getattr(s, 'shipment_direction', 'outgoing')
        ])
    return create_excel_file(data, headers, 'Pending Shipments', 'pending_shipments')

def export_customers(customers):
    headers = ['Customer Name', 'Email', 'Phone', 'City', 'State', 'Total Shipments', 'First Shipment', 'Last Shipment', 'Notes']
    data = []
    for c in customers:
        data.append([
            c.customer_name or '-',
            c.customer_email or '-',
            c.customer_phone or '-',
            c.city or '-',
            c.state or '-',
            c.total_shipments or 0,
            c.first_shipment.strftime('%Y-%m-%d') if c.first_shipment else '-',
            c.last_shipment.strftime('%Y-%m-%d') if c.last_shipment else '-',
            c.notes or '-'
        ])
    return create_excel_file(data, headers, 'Customers', 'customers')